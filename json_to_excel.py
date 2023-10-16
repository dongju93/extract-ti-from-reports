import json
import openpyxl
from envs.env import output_path

# Load the JSON data from a file
with open(output_path + "2012-2022_merged.json", "r", encoding="utf-8") as file:
    data = json.load(file)


def save_workbook(wb, count):
    """Save the current workbook with a specific name."""
    filename = output_path + f"excel/split_excel_{count}.xlsx"
    wb.save(filename)
    print(f"Saved: {filename}")


def sanitize_string(s):
    # For the purpose of the example, replace null characters.
    # You can extend this to handle other illegal characters.
    return s.replace("\u0000", "")


# Counters
sheet_count = 0
file_count = 1

# Create a new Excel workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

for item in data:
    # If we've hit 200 sheets, save current workbook and create a new one
    if sheet_count == 200:
        save_workbook(wb, file_count)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        sheet_count = 0
        file_count += 1

    rule_id = item["rule_id"]

    # Create a new sheet with the name of the rule_id
    ws = wb.create_sheet(title=str(rule_id))

    ws["A1"] = "rule_id"
    ws["A2"] = item["rule_id"]

    ws["B1"] = "name"
    ws["B2"] = item["name"]

    ws["C1"] = "description"
    ws["C2"] = item["description"]

    ws["D1"] = "samples"
    if isinstance(item["samples"], list):
        for index, sample in enumerate(item["samples"], start=2):
            ws.cell(row=index, column=4, value=sample)
    else:
        ws["D2"] = item["samples"]

    ws["E1"] = "weight"
    ws["E2"] = item["weight"]

    ws["F1"] = "signatures"
    if isinstance(item["signatures"], list):
        for index, signature in enumerate(item["signatures"], start=2):
            ws.cell(row=index, column=6, value=sanitize_string(signature))
    else:
        ws["F2"] = item["signatures"]

    sheet_count += 1

# Save any remaining workbook
if sheet_count > 0:
    save_workbook(wb, file_count)
